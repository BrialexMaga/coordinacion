from django.shortcuts import render, redirect
from .forms import byCycleForm, byCycleAndSubjectForm, byCycleRangeForm
from studentform.models import Subject
from studenthistory.models import Section
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import IntegerField
from django.db.models.functions import Cast

from django.contrib.auth.decorators import login_required

from django.db import connection

@login_required
def filterRateSubjectFailures(request):
    form = byCycleAndSubjectForm()
    form2 = byCycleForm()
    form3 = byCycleRangeForm()

    return render(request, 'rateSubjectFailures/failing_rate_filter.html', 
                  {'cycleRangeForm': form, 'cycleForm': form2, 'rangeForm': form3})

@login_required
def byCycleFilter(request):
    if request.method == 'POST':
        form = byCycleForm(request.POST)
        if form.is_valid():
            courses = form.filter_courses()
            statistics, registers = makeByCycleStatistics(courses)
            title = f'Ciclo {form.cleaned_data.get("school_cycle")}'

            # Save the information in session variables
            request.session['statistics'] = statistics
            request.session['registers'] = registers

            return render(request, 'rateSubjectFailures/by_cycle.html', 
                          {'statistics': statistics, 'registers': registers, 'title': title})

@login_required
def byCycleSubjectFilter(request):
    if request.method == 'POST':
        form = byCycleAndSubjectForm(request.POST)
        if form.is_valid():
            courses = form.filter_courses()
            title = f'{form.cleaned_data.get("school_cycle")} - {form.cleaned_data.get("subject")}'

            registers = makeByCycleSubjectRegisters(courses)

            # Save the information in session variables
            request.session['registers'] = registers
            request.session['title'] = title
            
            return render(request, 'rateSubjectFailures/by_cycle_subject.html', 
                          {'registers': registers, 'title': title, 'are_there_registers': len(registers) > 0})

@login_required
def byCycleRangeFilter(request):
    if request.method == 'POST':
        form = byCycleRangeForm(request.POST)
        if form.is_valid():
            courses = form.filter_courses()
            registers = makebyCycleRangeRegisters(courses)

            beginning = form.cleaned_data.get('start_cycle')
            end = form.cleaned_data.get('end_cycle')

            title = f'{beginning} - {end}'

            # Save the information in session variables
            request.session['registers'] = registers
            request.session['title'] = title

            return render(request, 'rateSubjectFailures/by_cycle_range.html', 
                          {'courses': courses, 'registers': registers, "title": title})






def makeByCycleStatistics(courses):
    subject_indexes = courses.values_list('section__subject__idSubject', flat=True).distinct()
    subjects = {subject.idSubject: subject for subject in Subject.objects.filter(idSubject__in=subject_indexes)}

    statistics = []
    registers = []
    for subject_index in subject_indexes:
        subject = subjects[subject_index]
        subject_courses = courses.filter(section__subject__idSubject=subject_index).order_by('section__section', 'student__idStudent', 
                                                                                                'grade_period__code_name')

        section_indexes = subject_courses.values_list('section__idSection', flat=True).distinct()
        section_indexes = list(set(section_indexes))
        sections = {section.idSection: section for section in Section.objects.filter(idSection__in=section_indexes)}
        
        section_fail_rates = []
        for section_index in section_indexes:
            calculatePercent = lambda x, y: round(x / y * 100, 2) if y else 0
            section_courses = subject_courses.filter(section__idSection=section_index)
            section = sections[section_index]

            total_students = section_courses.values_list('student__idStudent', flat=True).distinct().count()

            passed_extraordinary = section_courses.exclude(grade__in=['SD']).annotate(
                grade_int=Cast('grade', IntegerField())
            ).filter(grade_period__code_name='E', grade_int__gte=60, grade_int__lte=100).values_list('student__idStudent', flat=True).distinct()
            
            no_passed_extraordinary = len(passed_extraordinary)
            no_passed_ordinary = section_courses.exclude(grade__in=['SD']).annotate(
                grade_int=Cast('grade', IntegerField())
            ).filter(grade_period__code_name='OE', grade_int__gte=60, grade_int__lte=100).count()

            no_failed_extraordinary = section_courses.exclude(grade__in=['SD']).annotate(
                grade_int=Cast('grade', IntegerField())
            ).filter(grade_period__code_name='E', grade_int__lt=60).count()
            no_failed_extraordinary += section_courses.filter(grade_period__code_name='E', grade='SD').count()

            no_failed_ordinary = total_students - no_passed_ordinary

            section_courses_failed_numeric = section_courses.exclude(grade__in=['SD']).annotate(
                grade_int=Cast('grade', IntegerField())
            ).filter(grade_period__code_name='OE', grade_int__lt=60)
            section_courses_failed_numeric = section_courses_failed_numeric.exclude(student__idStudent__in=passed_extraordinary).count()

            section_courses_failed_sd = section_courses.filter(grade_period__code_name='OE', grade='SD')
            section_courses_failed_sd = section_courses_failed_sd.exclude(student__idStudent__in=passed_extraordinary).count()
            
            no_sd = section_courses_failed_sd + section_courses.filter(grade_period__code_name='E', grade='SD').count()

            section_courses_failed = section_courses_failed_numeric + section_courses_failed_sd

            SD_percent = calculatePercent(no_sd, total_students)
            OE_percent = calculatePercent(no_passed_ordinary, total_students)
            E_percent = calculatePercent(no_passed_extraordinary, total_students)
            NO_OE_percent = calculatePercent(no_failed_ordinary, total_students)
            NO_E_percent = calculatePercent(no_failed_extraordinary, total_students)

            failed_rate = section_courses_failed / total_students * 100
            failed_rate = round(failed_rate, 2)

            section_fail_rates.append(failed_rate)

            #print(len(connection.queries))
            registers.append({'key_subject': subject.key_subject, 'subject': subject.name, 'section': section.section, 'total_students': total_students,
                              'SD': no_sd, 'OE': no_passed_ordinary, 'E': no_passed_extraordinary, 'NO_OE': no_failed_ordinary,
                              'NO_E': no_failed_extraordinary, 'SD_percent': SD_percent, 'OE_percent': OE_percent, 'E_percent': E_percent,
                              'NO_OE_percent': NO_OE_percent, 'NO_E_percent': NO_E_percent})

        total_students = subject_courses.values_list('student__idStudent', flat=True).distinct().count()

        passed_extraordinary = subject_courses.exclude(grade__in=['SD']).annotate(
            grade_int=Cast('grade', IntegerField())
        ).filter(grade_period__code_name='E', grade_int__gte=60, grade_int__lte=100).values_list('student__idStudent', flat=True).distinct()

        subject_courses_failed_numeric = subject_courses.exclude(grade__in=['SD']).annotate(
                grade_int=Cast('grade', IntegerField())
            ).filter(grade_period__code_name='OE', grade_int__lt=60)
        subject_courses_failed_numeric = subject_courses_failed_numeric.exclude(student__idStudent__in=passed_extraordinary).count()

        subject_courses_failed_sd = subject_courses.filter(grade_period__code_name='OE', grade='SD')
        subject_courses_failed_sd = subject_courses_failed_sd.exclude(student__idStudent__in=passed_extraordinary).count()

        subject_courses_failed = subject_courses_failed_numeric + subject_courses_failed_sd

        failed_rate = subject_courses_failed / total_students * 100
        failed_rate = round(failed_rate, 2)

        section_fail_rate = sum(section_fail_rates) / len(section_fail_rates) if section_fail_rates else 0
        section_fail_rate = round(section_fail_rate, 2)

        statistics.append({'key_subject': subject.key_subject, 'subject': subject.name, 
                            'section_fail_rate': section_fail_rate, 'total_students': total_students, 
                            'total_failed': subject_courses_failed, 'failed_rate': failed_rate})
    
    return statistics, registers

def makeByCycleSubjectRegisters(courses):
    courses = courses.order_by('section__section', 'student__idStudent', 'grade_period__code_name')
    students = list(dict.fromkeys(courses.values_list('student__idStudent', flat=True).distinct()).keys())

    registers = []
    for student_id in students:
        course = courses.filter(student__idStudent=student_id)
        section = course.first().section
        student_info = course.first().student

        if course.count() > 1:
            grade_ordinary = course.filter(grade_period__code_name='OE').first().grade
            grade_extraordinary = course.filter(grade_period__code_name='E').first().grade
        else:
            grade_ordinary = course.first().grade
            grade_extraordinary = ""

        registers.append({"section": section.section, "name": f'{student_info.name} {student_info.first_last_name} {student_info.second_last_name}', 
                          "code": student_info.code, "OE": grade_ordinary, "E": grade_extraordinary})

    return registers

def makebyCycleRangeRegisters(courses):
    calculatePercent = lambda x, y: round(x / y * 100, 2) if y else 0
    subject_indexes = courses.values_list('section__subject__idSubject', flat=True).distinct()
    subjects = {subject.idSubject: subject for subject in Subject.objects.filter(idSubject__in=subject_indexes)}

    registers = []
    for subject_index in subject_indexes:
        subject = subjects[subject_index]
        subject_courses = courses.filter(section__subject__idSubject=subject_index).order_by('section__section', 'student__idStudent', 
                                                                                                'grade_period__code_name')
        
        total_students = subject_courses.values_list('student__idStudent', 'school_cycle__idCycle').distinct().count()

        passed_extraordinary = subject_courses.exclude(grade__in=['SD']).annotate(
            grade_int=Cast('grade', IntegerField())
        ).filter(grade_period__code_name='E', grade_int__gte=60, grade_int__lte=100).values_list('student__idStudent', flat=True).distinct()

        subject_courses_failed_numeric = subject_courses.exclude(grade__in=['SD']).annotate(
                grade_int=Cast('grade', IntegerField())
            ).filter(grade_period__code_name='OE', grade_int__lt=60)
        subject_courses_failed_numeric = subject_courses_failed_numeric.exclude(student__idStudent__in=passed_extraordinary).count()

        subject_courses_failed_sd = subject_courses.filter(grade_period__code_name='OE', grade='SD')
        subject_courses_failed_sd = subject_courses_failed_sd.exclude(student__idStudent__in=passed_extraordinary).count()

        subject_courses_failed = subject_courses_failed_numeric + subject_courses_failed_sd

        registers.append({'key_subject': subject.key_subject, 'subject': subject.name, 'total_students': total_students,
                          'failed_students': subject_courses_failed, 'failed_rate': calculatePercent(subject_courses_failed, total_students)})
    
    #print(len(connection.queries))
    return registers

@login_required
def exportByCycle(request):
    workbook = Workbook()
    registers_sheet = workbook.active
    registers_sheet.title = "Registros"

    # Add headers
    registers_data = [
        ["Clave", 
         "Materia", 
         "Sección", 
         "Total Alumnos", 
         "S/D", 
         "Acreditados Ordinario",
         "Acreditados Extraordinario",
         "No Acreditados Ordinario",
         "No Acreditados Extraordinario",
         "% S/D",
         "% Acreditados Ordinario",
         "% Acreditados Extraordinario",
         "% No Acreditados Ordinario",
         "% No Acreditados Extraordinario"],
    ]

    # Add data
    registers = request.session.get('registers', [])

    for register in registers:
        registers_data.append(
            [register['key_subject'], 
             register['subject'], 
             register['section'], 
             register['total_students'], 
             register['SD'], 
             register['OE'], 
             register['E'], 
             register['NO_OE'], 
             register['NO_E'], 
             register['SD_percent'], 
             register['OE_percent'], 
             register['E_percent'], 
             register['NO_OE_percent'], 
             register['NO_E_percent']]
        )

    # Add data to sheet
    for row in registers_data:
        registers_sheet.append(row)


    # Add statistics sheet
    statistics_sheet = workbook.create_sheet(title="Estadisticas")

    # Add headers
    statistics = request.session.get('statistics', {})
    statistics_data = [
        ["Clave",
         "Materia",
         "Promedio de reprobados por secciones",
         "Total de alumnos",
         "Total de alumnos reprobados",
         "Porcentaje de alumnos reprobados"],  # Header
    ]
    for statistics_info in statistics:
        statistics_data.append(
            [statistics_info['key_subject'],
             statistics_info['subject'],
             statistics_info['section_fail_rate'],
             statistics_info['total_students'],
             statistics_info['total_failed'],
             f'{statistics_info["failed_rate"]}%']
        )

    # Add data
    for row in statistics_data:
        statistics_sheet.append(row)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Indice de reprobados por ciclo.xlsx'
    workbook.save(response)

    return response

@login_required
def exportByCycleSubject(request):
    workbook = Workbook()
    registers_sheet = workbook.active
    registers_sheet.title = "Registros"

    # Add headers
    registers_data = [
        ["Sección",
         "Nombre",
         "Código",
         "Ordinario",
         "Extraordinario"],
    ]

    # Add data
    registers = request.session.get('registers', [])

    for register in registers:
        registers_data.append(
            [register['section'],
             register['name'],
             register['code'],
             register['OE'],
             register['E']]
        )

    # Add data to sheet
    for row in registers_data:
        registers_sheet.append(row)

    title = request.session.get('title', "No hay registros para mostrar") # Title has already been set in byCycleSubjectFilter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename={title}.xlsx"
    workbook.save(response)

    return response

@login_required
def exportByCycleRange(request):
    workbook = Workbook()
    registers_sheet = workbook.active
    registers_sheet.title = "Registros"

    # Add headers
    registers_data = [
        ["Clave",
         "Materia",
         "Total de alumnos",
         "Total de alumnos reprobados",
         "Porcentaje de alumnos reprobados"],
    ]

    # Add data
    registers = request.session.get('registers', [])

    for register in registers:
        registers_data.append(
            [register['key_subject'],
             register['subject'],
             register['total_students'],
             register['failed_students'],
             f'{register["failed_rate"]}%']
        )

    # Add data to sheet
    for row in registers_data:
        registers_sheet.append(row)

    title = request.session.get('title', "Sin registros")
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename={title}.xlsx"
    workbook.save(response)

    return response